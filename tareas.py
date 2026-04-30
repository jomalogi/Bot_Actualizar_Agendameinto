"""
Tarea Celery que ejecuta el agendamiento de órdenes en el módulo Claro.
Corre Selenium en modo headless (sin pantalla) para servidores Linux.
Emite progreso en tiempo real via Redis pub/sub → Socket.IO.
"""
import time
import logging
import os
from datetime import datetime
from celery import Celery
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.chrome.options import Options
from selenium.common.exceptions import TimeoutException, NoSuchElementException
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font

# ── Importar config ──
import sys
sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))
from config import Config

# ── Celery ──
celery = Celery("tareas", broker=Config.CELERY_BROKER, backend=Config.CELERY_BACKEND)

log = logging.getLogger("tarea_agendamiento")

# ─────────────────────────────────────────────────────
# CONSTANTES
# ─────────────────────────────────────────────────────
ESPERA_NORMAL  = Config.ESPERA_NORMAL
ESPERA_MODAL   = Config.ESPERA_MODAL
URL_BASE       = Config.URL_BASE
URL_AGENDAR    = Config.URL_AGENDAR
TIPO_DEFAULT   = "Orden de Trabajo"


# ─────────────────────────────────────────────────────
# UTILIDADES INTERNAS
# ─────────────────────────────────────────────────────
def _emit_progreso(task, trabajo_id: int, mensaje: str, procesadas: int,
                   total: int, estado_orden: dict = None):
    """Guarda estado en Celery meta para que Socket.IO lo retransmita."""
    task.update_state(
        state="PROGRESS",
        meta={
            "trabajo_id": trabajo_id,
            "mensaje":    mensaje,
            "procesadas": procesadas,
            "total":      total,
            "orden":      estado_orden,
        }
    )


def _iniciar_chrome_headless() -> webdriver.Chrome:
    import tempfile
    opts = Options()
    opts.add_argument("--headless=new")
    opts.add_argument("--no-sandbox")
    opts.add_argument("--disable-dev-shm-usage")
    opts.add_argument("--disable-gpu")
    opts.add_argument("--window-size=1920,1080")
    opts.add_argument("--disable-notifications")
    opts.add_argument("--log-level=3")
    opts.add_argument("--no-zygote")
    opts.add_argument("--disable-extensions")
    opts.add_argument("--disable-background-networking")
    opts.add_argument("--disable-setuid-sandbox")
    opts.add_argument("--ignore-certificate-errors")
    opts.add_argument("--ignore-ssl-errors")
    # Directorio temporal único por instancia para evitar conflictos
    user_data_dir = tempfile.mkdtemp(prefix="chrome_", dir="/tmp")
    opts.add_argument(f"--user-data-dir={user_data_dir}")

    try:
        from webdriver_manager.chrome import ChromeDriverManager
        driver = webdriver.Chrome(
            service=Service(ChromeDriverManager().install()), options=opts
        )
    except Exception:
        driver = webdriver.Chrome(options=opts)  # chromedriver en PATH
    return driver


def _esperar_pin_de_usuario(trabajo_id: int, timeout: int = 600) -> str | None:
    """Espera hasta timeout segundos a que el usuario ingrese el PIN via Redis."""
    import redis
    r = redis.from_url(Config.CELERY_BROKER)
    key = f"pin_trabajo_{trabajo_id}"
    r.delete(key)  # limpiar PIN anterior si existia

    inicio = time.time()
    while time.time() - inicio < timeout:
        try:
            val = r.get(key)
        except Exception:
            val = None
        if val:
            r.delete(key)
            return val.decode("utf-8").strip()
        time.sleep(2)
    return None


def _hacer_login(driver, wait, usuario_claro: str, password_claro: str,
                 trabajo_id: int = None, socketio=None):
    driver.get(URL_BASE)
    time.sleep(2)
    try:
        campo = wait.until(EC.presence_of_element_located(
            (By.XPATH, "//input[@type='text']")
        ))
        campo.clear()
        campo.send_keys(usuario_claro)
        driver.find_element(By.XPATH, "//input[@type='password']").send_keys(password_claro)
        driver.find_element(
            By.XPATH, "//input[@type='submit'] | //button[@type='submit']"
        ).click()
        time.sleep(3)
    except TimeoutException:
        pass

    # ── Detectar y manejar flujo 2FA (canalPin.php → validarPin.php) ──
    # Paso 1: seleccionar canal SMS si aparece canalPin
    if "canalPin" in driver.current_url:
        log.info("Pantalla canalPin detectada para trabajo %s", trabajo_id)
        try:
            radios = driver.find_elements(By.XPATH, "//input[@type='radio']")
            if radios:
                radios[0].click()
                time.sleep(0.5)
            driver.find_element(
                By.XPATH, "//input[@value='Enviar'] | //button[normalize-space()='Enviar']"
            ).click()
            time.sleep(3)
            log.info("Clic Enviar PIN realizado para trabajo %s", trabajo_id)
        except Exception as e:
            log.warning("Error en seleccion canal PIN: %s", e)

    # Paso 2: ingresar PIN si aparece validarPin
    if "validarPin" in driver.current_url:
        log.info("Pantalla validarPin detectada para trabajo %s", trabajo_id)

        # Notificar al frontend para que muestre el panel PIN
        if socketio and trabajo_id:
            socketio.emit("solicitar_pin", {
                "trabajo_id": trabajo_id,
                "mensaje": "Ingresa el PIN que recibiste por SMS."
            }, room=f"trabajo_{trabajo_id}")

        # Esperar PIN del usuario (max 10 minutos)
        pin = _esperar_pin_de_usuario(trabajo_id, timeout=600)

        if pin:
            try:
                campo_pin = WebDriverWait(driver, 15).until(
                    EC.presence_of_element_located(
                        (By.XPATH, "//input[@type='number'] | //input[@name='pin'] | //input[@type='text']")
                    )
                )
                campo_pin.clear()
                campo_pin.send_keys(pin)
                time.sleep(0.3)
                driver.find_element(
                    By.XPATH, "//button[normalize-space()='Validar'] | //input[@value='Validar']"
                ).click()
                time.sleep(3)
                log.info("PIN validado correctamente para trabajo %s", trabajo_id)
            except Exception as e:
                log.error("Error ingresando PIN: %s", e)
                raise Exception(f"Error al ingresar el PIN 2FA: {e}")
        else:
            raise Exception("Timeout esperando PIN del usuario (10 minutos)")


def _consultar_orden(driver, wait, numero_orden: str, tipo_orden: str):
    driver.get(URL_AGENDAR)
    time.sleep(1.5)
    campo = wait.until(EC.presence_of_element_located((By.XPATH, "//input[@type='text']")))
    campo.clear()
    campo.send_keys(str(numero_orden))

    if "edificio" in tipo_orden.lower():
        idx = 1
    elif "llamada" in tipo_orden.lower():
        idx = 2
    else:
        idx = 0

    radios = driver.find_elements(By.XPATH, "//input[@type='radio']")
    try:
        radio_txt = ["Orden de Trabajo", "Orden de Edificio", "Llamada de servicio"][idx]
        radio = driver.find_element(
            By.XPATH,
            f"//label[contains(text(),'{radio_txt}')]/preceding-sibling::input[@type='radio'] | "
            f"//label[contains(text(),'{radio_txt}')]/../input[@type='radio']"
        )
        if not radio.is_selected():
            radio.click()
    except NoSuchElementException:
        if radios and len(radios) > idx and not radios[idx].is_selected():
            radios[idx].click()

    time.sleep(0.5)
    wait.until(EC.element_to_be_clickable(
        (By.XPATH, "//input[@value='Consultar'] | //button[normalize-space()='Consultar']")
    )).click()
    time.sleep(2)


def _verificar_error_post_consulta(driver) -> str | None:
    try:
        el = driver.find_element(
            By.XPATH,
            "//*[contains(text(),'no se puede agendar') or contains(text(),'no contiene') "
            "or contains(text(),'cerrada en RR') or contains(text(),'Orden cerrada')]"
        )
        return el.text.strip()
    except NoSuchElementException:
        return None


def _clic_agendar(driver, wait):
    """Detecta Agendar o Re Agendar y actúa en consecuencia."""
    xpath_ag  = "//input[@value='Agendar']    | //button[normalize-space()='Agendar']"
    xpath_rag = "//input[@value='Re Agendar'] | //button[contains(text(),'Re Agendar')]"

    wait.until(EC.presence_of_element_located(
        (By.XPATH, f"{xpath_ag} | {xpath_rag}")
    ))

    es_re = bool(driver.find_elements(By.XPATH, xpath_rag))

    if es_re:
        driver.find_elements(By.XPATH, xpath_rag)[0].click()
        time.sleep(2)
        _confirmar_modal_reagendar(driver)
    else:
        driver.find_elements(By.XPATH, xpath_ag)[0].click()
        time.sleep(1.5)

    return es_re


def _confirmar_modal_reagendar(driver):
    """Llena observación y confirma el modal de Re Agendamiento."""
    try:
        WebDriverWait(driver, 5).until(EC.visibility_of_element_located(
            (By.XPATH, "//*[contains(text(),'Motivo de Re Agendamiento')]")
        ))
    except TimeoutException:
        pass
    time.sleep(0.8)

    try:
        ta = WebDriverWait(driver, 5).until(EC.element_to_be_clickable(
            (By.XPATH, "//label[contains(text(),'Observaci')]/following::textarea[1]")
        ))
    except TimeoutException:
        tas = driver.find_elements(By.XPATH, "//textarea")
        ta = next((t for t in tas if t.is_displayed() and t.is_enabled()), None)
        if not ta:
            return

    driver.execute_script("arguments[0].click();", ta)
    driver.execute_script("arguments[0].value='Agendar';", ta)
    driver.execute_script(
        "arguments[0].dispatchEvent(new Event('input',{bubbles:true}));", ta
    )
    time.sleep(0.5)

    btn = WebDriverWait(driver, 8).until(EC.element_to_be_clickable(
        (By.XPATH, "//button[normalize-space()='Confirmar'] | //input[@value='Confirmar']")
    ))
    driver.execute_script("arguments[0].click();", btn)

    # Esperar que el modal desaparezca
    try:
        WebDriverWait(driver, 10).until(EC.staleness_of(btn))
    except TimeoutException:
        pass
    time.sleep(2)


def _seleccionar_cupo(driver, wait) -> bool:
    """Selecciona el día más cercano con cupos disponibles (verde)."""
    wait.until(EC.presence_of_element_located((By.XPATH, "//table")))
    time.sleep(1.5)

    for _ in range(4):
        # Buscar TODAS las celdas td que contengan la palabra 'cupos' (case-insensitive)
        celdas = driver.find_elements(By.XPATH, "//td")
        disponibles = []
        for c in celdas:
            try:
                texto = c.text.strip().lower()
                # Debe contener 'cupos' y no ser '0 cupos'
                if 'cupos' not in texto:
                    continue
                # Extraer el número de cupos
                partes = texto.replace('\n', ' ').split()
                numero = None
                for p in partes:
                    if p.isdigit():
                        numero = int(p)
                        break
                    # Puede venir como '2cupos' pegado
                    digits = ''.join(filter(str.isdigit, p))
                    if digits:
                        numero = int(digits)
                        break
                if numero is None or numero == 0:
                    continue
                # Verificar que la celda esté visible y sea clickeable
                if not c.is_displayed():
                    continue
                # Verificar color de fondo (verde = disponible)
                bg = driver.execute_script(
                    "return window.getComputedStyle(arguments[0]).backgroundColor;", c
                )
                # Aceptar verde (rgb con G alto) O cualquier celda con cupos > 0
                # La validación de cupos > 0 ya es suficiente
                disponibles.append((numero, c))
            except Exception:
                continue

        if disponibles:
            # Ordenar por posición en pantalla (más cercano = más arriba/izquierda)
            disponibles.sort(key=lambda x: x[0], reverse=False)  # menor cupo primero no
            # Tomar el primero en orden DOM (ya está ordenado por aparición)
            celda_elegida = disponibles[0][1]
            log.info("Cupo encontrado: %s cupos - texto: %s",
                     disponibles[0][0], celda_elegida.text.strip()[:30])
            driver.execute_script("arguments[0].scrollIntoView({block:'center'});", celda_elegida)
            time.sleep(0.5)
            driver.execute_script("arguments[0].click();", celda_elegida)
            time.sleep(1.5)
            return True

        # No hay cupos en esta semana, intentar siguiente
        try:
            sig = driver.find_element(
                By.XPATH,
                "//a[contains(text(),'►')] | //input[contains(@value,'►')] | "
                "//a[contains(@onclick,'siguiente')] | //img[contains(@src,'next')]/parent::a"
            )
            sig.click()
            time.sleep(2)
        except NoSuchElementException:
            break
    return False


def _confirmar_agendamiento(driver) -> tuple[bool, str]:
    """
    Espera el modal 'Confirmar información' y hace clic en Confirmar.
    Detecta también errores (Advertencia, Orden cerrada).
    """
    wait_m = WebDriverWait(driver, ESPERA_MODAL)

    xpath_ok     = "//*[contains(text(),'Confirmar informaci')]"
    xpath_err    = "//*[contains(text(),'Advertencia')]"
    xpath_closed = "//*[contains(text(),'Orden cerrada') or contains(text(),'cerrada en RR')]"
    xpath_acept  = "//button[normalize-space()='Aceptar'] | //input[@value='Aceptar']"

    try:
        wait_m.until(lambda d:
            d.find_elements(By.XPATH, xpath_ok) or
            d.find_elements(By.XPATH, xpath_err) or
            d.find_elements(By.XPATH, xpath_closed)
        )
    except TimeoutException:
        return False, "Timeout esperando modal de confirmación"

    # Orden cerrada
    if driver.find_elements(By.XPATH, xpath_closed):
        btns = driver.find_elements(By.XPATH, xpath_acept)
        if btns:
            driver.execute_script("arguments[0].click();", btns[0])
        time.sleep(1)
        return False, "Orden Cerrada en RR"

    # Error / Advertencia
    if driver.find_elements(By.XPATH, xpath_err):
        try:
            msg = driver.find_element(
                By.XPATH, xpath_err + "/ancestor::*[2]"
            ).text.strip()
        except Exception:
            msg = "Error/Advertencia"
        btns = driver.find_elements(By.XPATH, xpath_acept)
        if btns:
            driver.execute_script("arguments[0].click();", btns[0])
        time.sleep(1)
        return False, msg

    # Modal de confirmación OK → clic en Confirmar
    xpath_btn = (
        "//*[contains(text(),'Confirmar informaci')]"
        "/ancestor::*[position()<=6]"
        "//button[normalize-space()='Confirmar'] | "
        "//*[contains(text(),'Confirmar informaci')]"
        "/ancestor::*[position()<=6]//input[@value='Confirmar']"
    )
    try:
        btn = WebDriverWait(driver, 10).until(
            EC.element_to_be_clickable((By.XPATH, xpath_btn))
        )
    except TimeoutException:
        btns = [b for b in driver.find_elements(
            By.XPATH, "//button[normalize-space()='Confirmar']"
        ) if b.is_displayed() and b.is_enabled()]
        if not btns:
            return False, "No se encontró botón Confirmar"
        btn = btns[0]

    driver.execute_script("arguments[0].scrollIntoView({block:'center'});", btn)
    time.sleep(0.5)
    driver.execute_script("arguments[0].click();", btn)

    # Esperar mensaje de éxito
    xpath_exito = (
        "//*[contains(text(),'realizada correctamente') or "
        "contains(text(),'realizado correctamente') or "
        "contains(text(),'Reagenda realizada') or "
        "contains(text(),'Agenda realizada')]"
    )
    try:
        WebDriverWait(driver, ESPERA_MODAL).until(
            EC.visibility_of_element_located((By.XPATH, xpath_exito))
        )
        btns_ok = driver.find_elements(By.XPATH, xpath_acept)
        if btns_ok:
            driver.execute_script("arguments[0].click();", btns_ok[0])
        time.sleep(1)
        return True, ""
    except TimeoutException:
        # Verificar estado de visita como fallback
        for est in ["REPROGRAMADA", "PROGRAMADA", "AGENDADO"]:
            if driver.find_elements(By.XPATH, f"//*[contains(text(),'{est}')]"):
                return True, ""
        return True, "Verificar manualmente"


def _guardar_excel_resultado(ruta_entrada: str, ruta_salida: str, resultados: list[dict]):
    """Escribe los resultados en el Excel y lo guarda en ruta_salida."""
    wb = load_workbook(ruta_entrada)
    ws = wb.active

    enc = [c.value for c in ws[1]]
    for col_name in ["Estado", "Detalle", "Procesado_en"]:
        if col_name not in enc:
            ws.cell(row=1, column=len(enc) + 1, value=col_name)
            enc.append(col_name)

    col_estado = enc.index("Estado") + 1
    col_det    = enc.index("Detalle") + 1
    col_fecha  = enc.index("Procesado_en") + 1

    colores = {
        "AGENDADO":           "C6EFCE",
        "RE-AGENDADO":        "BDD7EE",
        "Orden Cerrada en RR":"FFEB9C",
        "ERROR":              "FFC7CE",
    }

    for res in resultados:
        fila  = res["fila"]
        est   = res["estado"]
        color = colores.get(est, "FFC7CE")
        fill  = PatternFill(start_color=color, end_color=color, fill_type="solid")

        c1 = ws.cell(row=fila, column=col_estado, value=est)
        c2 = ws.cell(row=fila, column=col_det,    value=res.get("detalle", ""))
        c3 = ws.cell(row=fila, column=col_fecha,  value=res.get("fecha", ""))
        for c in (c1, c2, c3):
            c.fill = fill
        c1.font = Font(bold=True)

    wb.save(ruta_salida)


# ─────────────────────────────────────────────────────
# TAREA PRINCIPAL CELERY
# ─────────────────────────────────────────────────────
@celery.task(bind=True, name="tareas.ejecutar_agendamiento")
def ejecutar_agendamiento(self, trabajo_id: int, ruta_excel: str,
                           usuario_claro: str, password_claro: str):
    """
    Tarea principal: procesa todas las órdenes del Excel
    y emite progreso en tiempo real.
    """
    from app import app, db, socketio
    from models import Trabajo, OrdenResultado

    resultados = []
    driver = None

    with app.app_context():
        db.session.rollback()
        trabajo = db.session.get(Trabajo, trabajo_id)
        trabajo.estado = "ejecutando"
        db.session.commit()

        try:
            # ── Leer Excel ──
            wb    = load_workbook(ruta_excel)
            ws    = wb.active
            enc   = [c.value for c in ws[1]]
            filas = []
            for idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
                if not row[0]:
                    continue
                o = dict(zip(enc, row))
                o["_fila"] = idx
                # Saltar ya procesadas
                if o.get("Estado") in ("AGENDADO", "RE-AGENDADO"):
                    continue
                filas.append(o)

            trabajo.total = len(filas)
            db.session.commit()

            def emit(msg, procesadas, orden_info=None):
                _emit_progreso(self, trabajo_id, msg, procesadas, trabajo.total, orden_info)
                socketio.emit("progreso", {
                    "trabajo_id": trabajo_id,
                    "mensaje":    msg,
                    "procesadas": procesadas,
                    "total":      trabajo.total,
                    "exitosas":   trabajo.exitosas,
                    "fallidas":   trabajo.fallidas,
                    "orden":      orden_info,
                }, room=f"trabajo_{trabajo_id}")

            emit(f"Iniciando — {trabajo.total} órdenes pendientes", 0)

            # ── Iniciar Chrome headless ──
            driver = _iniciar_chrome_headless()
            wait   = WebDriverWait(driver, ESPERA_NORMAL)
            _hacer_login(driver, wait, usuario_claro, password_claro,
                         trabajo_id=trabajo_id, socketio=socketio)
            emit("Login completado ✅", 0)

            REINICIAR_CADA = 999999  # reinicio desactivado

            # ── Procesar órdenes ──
            import redis as _redis
            _r = _redis.from_url(Config.CELERY_BROKER)

            for i, orden in enumerate(filas, start=1):
                num   = str(orden.get("numero_orden", "")).strip()
                tipo  = str(orden.get("tipo_orden", TIPO_DEFAULT)).strip()
                fila  = orden["_fila"]

                # ── Reiniciar Chrome cada N órdenes para evitar crashes de memoria ──
                if i > 1 and (i - 1) % REINICIAR_CADA == 0:
                    emit(f"🔄 Reiniciando navegador para liberar memoria ({i-1}/{trabajo.total})...", i - 1)
                    log.info("Reiniciando Chrome en orden %s/%s", i, trabajo.total)
                    try:
                        driver.quit()
                    except Exception:
                        pass
                    time.sleep(2)
                    driver = _iniciar_chrome_headless()
                    wait   = WebDriverWait(driver, ESPERA_NORMAL)
                    _hacer_login(driver, wait, usuario_claro, password_claro,
                                 trabajo_id=trabajo_id, socketio=socketio)
                    emit(f"✅ Navegador reiniciado, continuando...", i - 1)

                emit(f"Procesando orden {i}/{trabajo.total}: {num}", i - 1,
                     {"numero": num, "estado": "procesando"})

                estado_final = "ERROR"
                detalle      = ""
                if _r.get(f"detener_trabajo_{trabajo_id}"):
                    log.info("Trabajo %s detenido por el usuario", trabajo_id)
                    emit(f"⏹ Proceso detenido por el usuario ({i-1}/{trabajo.total} procesadas)", i - 1)
                    break

                try:
                    _consultar_orden(driver, wait, num, tipo)

                    # Error post-consulta
                    err = _verificar_error_post_consulta(driver)
                    if err:
                        if "cerrada" in err.lower():
                            estado_final = "Orden Cerrada en RR"
                        else:
                            estado_final, detalle = "ERROR", err
                        raise StopIteration

                    # Verificar error inmediato antes del calendario
                    try:
                        err_elem = WebDriverWait(driver, 5).until(
                            EC.visibility_of_element_located((By.XPATH,
                                "//*[contains(text(),'Advertencia') or "
                                "contains(text(),'no contiene') or "
                                "contains(text(),'no se puede agendar') or "
                                "contains(text(),'cerrada en RR')]"
                            ))
                        )
                        msg_err = err_elem.text.strip()
                        if "cerrada" in msg_err.lower():
                            estado_final = "Orden Cerrada en RR"
                        else:
                            estado_final, detalle = "ERROR", msg_err
                        try:
                            driver.find_element(
                                By.XPATH,
                                "//button[normalize-space()='Aceptar'] | //input[@value='Aceptar']"
                            ).click()
                        except Exception:
                            pass
                        raise StopIteration
                    except TimeoutException:
                        pass

                    es_re = _clic_agendar(driver, wait)

                    hay_cupos = _seleccionar_cupo(driver, wait)
                    if not hay_cupos:
                        estado_final = "ERROR"
                        detalle      = "Sin cupos disponibles"
                        raise StopIteration

                    ok, det = _confirmar_agendamiento(driver)
                    if ok:
                        estado_final = "RE-AGENDADO" if es_re else "AGENDADO"
                        detalle      = det
                    else:
                        estado_final = det if det in ("Orden Cerrada en RR",) else "ERROR"
                        detalle      = det

                except StopIteration:
                    pass
                except Exception as ex:
                    estado_final = "ERROR"
                    detalle      = str(ex)[:200]

                fecha_ahora = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
                resultados.append({
                    "fila": fila, "estado": estado_final,
                    "detalle": detalle, "fecha": fecha_ahora
                })

                # Guardar en BD
                db.session.rollback()
                or_ = OrdenResultado(
                    trabajo_id=trabajo_id, numero_orden=num,
                    tipo_orden=tipo, estado=estado_final, detalle=detalle
                )
                db.session.add(or_)
                trabajo = db.session.get(Trabajo, trabajo_id)
                trabajo.procesadas = i
                if estado_final in ("AGENDADO", "RE-AGENDADO"):
                    trabajo.exitosas += 1
                else:
                    trabajo.fallidas += 1
                db.session.commit()

                emit(f"Orden {num} → {estado_final}", i,
                     {"numero": num, "estado": estado_final, "detalle": detalle})

                # ── Verificar señal de detener post-orden ──
                if _r.get(f"detener_trabajo_{trabajo_id}"):
                    log.info("Trabajo %s detenido por el usuario post-orden", trabajo_id)
                    emit(f"⏹ Proceso detenido por el usuario ({i}/{trabajo.total} procesadas)", i)
                    break

                time.sleep(1.5)

            # ── Guardar Excel resultado ──
            nombre_salida  = f"resultado_{trabajo_id}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
            ruta_salida    = os.path.join(Config.OUTPUT_FOLDER, nombre_salida)
            _guardar_excel_resultado(ruta_excel, ruta_salida, resultados)

            db.session.rollback()
            trabajo = db.session.get(Trabajo, trabajo_id)
            trabajo.estado         = "completado"
            trabajo.archivo_salida = nombre_salida
            trabajo.terminado_en   = datetime.now()
            db.session.commit()

            emit(f"✅ Completado — {trabajo.exitosas} exitosas, {trabajo.fallidas} con error",
                 trabajo.total)

        except Exception as ex:
            if trabajo:
                try:
                    db.session.rollback()
                    trabajo = db.session.get(Trabajo, trabajo_id)
                    trabajo.estado = "error"
                    db.session.commit()
                except Exception:
                    pass
            emit(f"❌ Error general: {str(ex)[:100]}", trabajo.procesadas if trabajo else 0)
            raise

        finally:
            if driver:
                try:
                    driver.quit()
                except Exception:
                    pass

    return {"trabajo_id": trabajo_id, "estado": "completado"}
